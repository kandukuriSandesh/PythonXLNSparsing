import openpyxl as xl
from openpyxl.chart import Reference,BarChart

def process_spreadsheet(filename):
    wb = xl.load_workbook(filename)
    sheet = wb['Sheet1']

    for row in range(2,sheet.max_row + 1):
        cell = sheet.cell(row,3)
        cell.value = cell.value * 0.9
        print(cell.value)
      
    values = Reference(sheet,min_row =2,max_row = sheet.max_row,min_col=4,max_col=4)
    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart,'e2')
    wb.save(f'transactions2.xlsx')
      

process_spreadsheet('transactions.xlsx')